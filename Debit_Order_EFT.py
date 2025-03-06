import tkinter as tk
from tkinter import *
from tkinter import Tk, Button, Label, filedialog, messagebox, StringVar
from PIL import Image, ImageTk
import sqlite3
import pandas as pd
import shutil
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl.styles.numbers import FORMAT_NUMBER_00  # Format for 2 decimal places

# Global data frames (these need to be populated by load functions)
eft_file_df = []  # Placeholder for the eft_file_df (to be populated from the .eft file)
billing_df = []  # Placeholder for the billing_df (to be populated from the CSV)
updated_df = []  # Placeholder for the updated DataFrame

def update_status(label_var, label_widget, status):
    """Update the status message for a specific process and set the color."""
    label_var.set(status)
    if status == "Complete":
        label_widget.config(fg="#00FF00")  # Green for "Complete"
    else:
        label_widget.config(fg="#FF0000")  # Red for "Not processed"

# Function to handle the rounding logic
def round_amount(amount):
    # Round the amount according to the specified rules
    amount = int(amount)
    
    # Get the last digit of the amount
    last_digit = amount % 10
    
    # If the last digit is one of the specified values, round accordingly
    if last_digit in {4, 14, 24, 34, 44, 54, 64, 74, 84, 94}:
        # Round to the next multiple of 5
        amount = (amount // 10) * 10 + 5
    elif last_digit in {9, 19, 29, 39, 49, 59, 69, 79, 89, 99}:
        # Round to the next multiple of 10
        amount = (amount // 10) * 10 + 10

    # Return the amount padded to 11 digits
    return f"{amount:011d}"

# Load CSV file and process the DataFrame
def load_csv_file():
    """Function to load the CSV file and process it as per the instructions."""
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    
    if not file_path:
        return  # If no file is selected, exit the function

    # Load the CSV into a DataFrame
    global billing_df
    billing_df = pd.read_csv(file_path)

    # Consolidate data by 'SabreCode' and calculate the sum of 'TotalDue'
    billing_df = billing_df.groupby('SabreCode', as_index=False)['TotalDue'].sum()

    # Format 'SabreCode' to have a leading zero, ensuring it's 5 characters
    billing_df['SabreCode'] = billing_df['SabreCode'].apply(lambda x: f"{str(x).zfill(7)}")

    # Adjust 'TotalDue' column (multiply by 1.15 and then by 100)
    billing_df['TotalDue'] = billing_df['TotalDue'] * 1.15 * 100

    # Round the 'TotalDue' values
    billing_df['TotalDue'] = billing_df['TotalDue'].apply(round_amount)

    # Show a message box confirming the CSV data import
    messagebox.showinfo("Success", "CSV data imported successfully!")
    update_status(csv_status, csv_status_label, "Complete")    

    # Display the updated DataFrame for debugging purposes
    print(billing_df)

    # Update the status label to reflect the successful load
    update_status(csv_status, csv_status_label, "Complete")

# Load EFT file function
def load_eft_file():
    """
    Function to load an .eft file, process it into a DataFrame, and update status indicators.
    """
    global eft_file_df, column_headings  # Declare global variables for the DataFrame and column headings
    
    # Prompt user to select an .eft file
    file_path = filedialog.askopenfilename(title="Open .eft File", filetypes=(("Text files", "*.eft"), ("All files", "*.*")))
    if not file_path:
        return  # Exit if no file is selected

    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        # Skip the first line (header) and process the remaining lines
        data_lines = lines[1:]

        # Process each line: split by '  ' (two spaces) and remove extra spaces
        processed_data = []
        column_counts = set()  # To track unique column counts
        max_columns = 0  # To keep track of the maximum number of columns

        for line_num, line in enumerate(data_lines, start=2):  # start=2 to account for skipping the first line
            line = line.strip()
            if not line:  # Skip empty lines
                continue
            split_line = [item.strip() for item in line.split('  ') if item.strip()]
            processed_data.append(split_line)
            column_counts.add(len(split_line))
            max_columns = max(max_columns, len(split_line))  # Update maximum column count

        # Normalize the rows to match the maximum number of columns (pad shorter rows)
        for i in range(len(processed_data)):
            current_length = len(processed_data[i])
            if current_length < max_columns:
                processed_data[i].extend([''] * (max_columns - current_length))  # Pad with empty strings
            elif current_length > max_columns:
                processed_data[i] = processed_data[i][:max_columns]  # Trim extra columns

        # Generate column headings dynamically
        column_headings = [f"Column {i+1}" for i in range(max_columns)]

        # Rename specific columns based on position
        if max_columns >= 1: column_headings[0] = "SabreCode"
        if max_columns >= 4: column_headings[3] = "BranchCode"
        if max_columns >= 5: column_headings[4] = "AccNumber"
        if max_columns >= 6: column_headings[5] = "CompanyName"
        if max_columns >= 7: column_headings[6] = "TotalDue"

        # Create a DataFrame with the processed data and column headings
        eft_file_df = pd.DataFrame(processed_data, columns=column_headings)

        # Show a success message and update status
        messagebox.showinfo("Success", "EFT File imported successfully!")
        update_status(eft_status, eft_status_label, "Complete")

        # Optionally print a preview of the DataFrame for debugging
        print(eft_file_df.head())

    except Exception as e:
        # Show an error message in case of failure
        messagebox.showerror("Error", f"An error occurred while processing the EFT file: {str(e)}")
        update_status(eft_status, eft_status_label, "Failed")

# Update Data function
def update_data():
    """
    Function to create 'updated_df' by copying 'eft_file_df' and updating the 'TotalDue'
    using values from 'billing_df' matching on 'SabreCode'. If no match exists, 'TotalDue' is set to 0.
    """
    global eft_file_df, billing_df, updated_df
    
    if eft_file_df is None or billing_df is None:
        messagebox.showerror("Error", "Please load both the EFT and Billing files before updating data.")
        return
    
    try:
        # Copy eft_file_df to create updated_df
        updated_df = eft_file_df.copy()

        # Update 'TotalDue' in updated_df by matching 'SabreCode' from billing_df
        updated_df = updated_df.merge(billing_df[['SabreCode', 'TotalDue']], on='SabreCode', how='left', suffixes=('', '_billing'))

        # If 'TotalDue_billing' is NaN, it means there was no match, so set 'TotalDue' to 0
        updated_df['TotalDue'] = updated_df['TotalDue_billing'].fillna(0)  # Replace NaN with 0 where no match was found
        
        # Drop the 'TotalDue_billing' column which we no longer need
        updated_df.drop(columns=['TotalDue_billing'], inplace=True)

        # Show a success message
        messagebox.showinfo("Info", "Updated Data created successfully!")
        update_status(updated_status, updated_status_label, "Complete")

        # Print updated_df for debugging to ensure it's correct
        print("Updated Data:")
        print(updated_df.head())

        # Optionally, if you want to display the first few rows in the Tkinter window, you can update the GUI text here.

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while updating the data: {str(e)}")
        update_status(updated_status, updated_status_label, "Failed")

# Export to Excel file function
def export_to_excel(eft_file_df, updated_df):
    # Prompt the user for the file save location
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

    if not file_path:
        return  # If no file path is selected, do nothing

    try:
        # Prepare the data for export
        updated_df_numeric = updated_df.copy()
        # Convert "TotalDue" to numeric and divide by 100
        updated_df_numeric["TotalDue"] = pd.to_numeric(updated_df_numeric["TotalDue"], errors="coerce") / 100

        # Ensure unique index for mapping (based on "SabreCode")
        updated_df_unique = updated_df_numeric.drop_duplicates(subset="SabreCode")

        # Create the export DataFrame by selecting necessary columns from eft_file_df
        export_df = eft_file_df[["SabreCode", "BranchCode", "AccNumber", "CompanyName"]].copy()

        # Map the "TotalDue" from updated_df to export_df based on "SabreCode"
        export_df["TotalDue"] = export_df["SabreCode"].map(updated_df_unique.set_index("SabreCode")["TotalDue"])

        # Ensure "eft_file_df" has unique "SabreCode" values for reindexing
        eft_file_unique = eft_file_df.drop_duplicates(subset="SabreCode")

        # Map the "TotalDue" from eft_file_df to "PrevMonthTotalDue"
        export_df["PrevMonthTotalDue"] = export_df["SabreCode"].map(eft_file_unique.set_index("SabreCode")["TotalDue"])

        # Ensure that both "TotalDue" and "PrevMonthTotalDue" are numeric (float) values
        export_df["TotalDue"] = pd.to_numeric(export_df["TotalDue"], errors="coerce")
        export_df["PrevMonthTotalDue"] = pd.to_numeric(export_df["PrevMonthTotalDue"], errors="coerce") / 100

        # Calculate the "Difference" column
        export_df["Difference"] = export_df["TotalDue"] - export_df["PrevMonthTotalDue"]

        # Now we should have exactly 7 columns: "SabreCode", "BranchCode", "AccNumber", "CompanyName", "TotalDue", "PrevMonthTotalDue", "Difference"
        column_headings = ["SabreCode", "BranchCode", "AccNumber", "CompanyName", "TotalDue", "PrevMonthTotalDue", "Difference"]

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Exported Data"

        # Write the column headings with formatting
        fill_color = PatternFill(start_color="CAF2F0", end_color="CAF2F0", fill_type="solid")
        for col_num, heading in enumerate(column_headings, start=1):
            cell = ws.cell(row=1, column=col_num, value=heading)
            cell.fill = fill_color
            cell.font = Font(bold=True)

        # Write the data to the sheet
        for row_num, row_data in enumerate(export_df.itertuples(index=False), start=2):
            for col_num, (col_name, cell_value) in enumerate(zip(export_df.columns, row_data), start=1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)

                # Format numeric columns
                if col_name in ["TotalDue", "PrevMonthTotalDue", "Difference"]:
                    cell.number_format = FORMAT_NUMBER_00

                # Apply conditional formatting to the "Difference" column
                if col_name == "Difference":
                    if cell_value < 0:
                        cell.font = Font(color="FF0000")  # Red for negative values
                    elif cell_value > 0:
                        cell.font = Font(color="0000FF")  # Blue for positive values

        # Save the workbook
        wb.save(file_path)

        # Show a success message
        messagebox.showinfo("Success", "Data exported successfully!")
        update_status(export_status, export_status_label, "Complete")

    except Exception as e:
        # Show an error message if an exception occurs
        messagebox.showerror("Error", f"An error occurred while exporting: {str(e)}")

# Create new EFT file function
def create_new_eft_file():
    global updated_df
    
    # Ask the user to save the new EFT file
    save_path = filedialog.asksaveasfilename(title="Save New EFT File", defaultextension=".eft", filetypes=(("Text files", "*.eft"), ("All files", "*.*")))
    if not save_path:
        return

    try:
        # Ask the user to select the original .eft file
        original_file_path = filedialog.askopenfilename(title="Open Original .eft File", filetypes=(("Text files", "*.eft"), ("All files", "*.*")))
        if not original_file_path:
            return
        
        # Read the original .eft file to preserve the first line (header) and track spacing
        with open(original_file_path, 'r', encoding='utf-8') as file:
            original_lines = file.readlines()

        header_line = original_lines[0]  # Preserve the first line (header)

        # Process the lines to preserve spacing and split by double spaces '  '
        spacing_info = []  # Will hold the space count between columns in each row
        data_lines = original_lines[1:]  # Skip header

        for line in data_lines:
            line = line.rstrip()  # Remove trailing spaces
            if not line:  # Skip empty lines
                continue

            # Regex to capture all column separators (two spaces or more) between data
            columns = re.split(r'( {2,})', line)  # Split on two or more spaces but keep the spaces
            spacing_info.append(columns)  # Save the line with spacing

        # Create the new EFT file and write the updated data with proper spacing
        with open(save_path, 'w', encoding='utf-8') as new_file:
            new_file.write(header_line)  # Write the header line first

            # Write each row from the updated data
            for idx, row in updated_df.iterrows():
                row_str = ''
                for i, col in enumerate(row):
                    # Handle TotalDue padding for 0 values
                    if isinstance(col, (int, float)) and col == 0:
                        col = '00000000000'  # Pad with 11 zeros if TotalDue is 0

                    # Add the correct spacing between columns from the original file
                    try:
                        if i > 0:  # For every column except the first one
                            space = spacing_info[idx][i * 2 - 1]  # Spacing is stored in between each column
                            row_str += space  # Add the space between columns
                        row_str += str(col)  # Add the column data
                    except IndexError as e:
                        print(f"IndexError at row {idx}, column {i}: {e}")
                        print(f"spacing_info[{idx}]: {spacing_info[idx]}")
                        print(f"row: {row}")
                        print(f"spacing_info length: {len(spacing_info)}")
                        print(f"row length: {len(row)}")
                        raise

                new_file.write(row_str + '\n')

        # Show a success message
        messagebox.showinfo("Success", "New EFT file created successfully!")
        update_status(eft_creation_status, eft_creation_status_label, "Complete")

    except Exception as e:
        # Show an error message if an exception occurs
        messagebox.showerror("Error", f"An error occurred while creating the new EFT file: {str(e)}")

# Create the GUI window
root = Tk()
root.title("Debit Order Updater")

# Set the window size (width x height)
root.geometry("400x500")

# Set the background color of the window
root.configure(bg="white")

# Set the window icon (make sure the file exists in your project directory)
#root.iconbitmap(r"c:\Users\ryadya\Conda\Scripts\Debit Order\bank_78392.ico")

# Load and display the logo
try:
    # Load and resize the logo image
    logo_image = Image.open(r"C:\Users\ryadya\Conda\Scripts\DebitOrder\Final\bank.png")
    logo_image = logo_image.resize((60, 60), Image.Resampling.LANCZOS)  # Use Resampling.LANCZOS directly
    logo_photo = ImageTk.PhotoImage(logo_image)

    # Create a label to display the logo and add it to the GUI
    logo_label = tk.Label(root, image=logo_photo, bg="white")  # Adjust background to match GUI
    logo_label.image = logo_photo  # Keep a reference to avoid garbage collection
    logo_label.pack(pady=(10, 30))
except Exception as e:
    messagebox.showerror("Error", f"Unable to load logo: {str(e)}")

# Status message variables
csv_status = StringVar(value="Not processed")
eft_status = StringVar(value="Not processed")
updated_status = StringVar(value="Not processed")
export_status = StringVar(value="Not processed")
eft_creation_status = StringVar(value="Not processed")

# Create a button to load the CSV file and its status label
load_csv_button = Button(root, text="Load Bill Run CSV File", command=load_csv_file, bg="#009688", fg="white")
load_csv_button.pack(ipadx=25, pady=10)
csv_status_label = Label(root, textvariable=csv_status, bg="white", fg="#FF0000")
csv_status_label.pack(pady=2)

# Create a button to load the .eft file and its status label
load_button = Button(root, text="Load Prev. Month .eft File", command=load_eft_file, bg="#CCECFF", fg="black") # Prev color #99CCFF
load_button.pack(ipadx=16, pady=10)
eft_status_label = Label(root, textvariable=eft_status, bg="white", fg="#FF0000")
eft_status_label.pack(pady=2)

# Button to update data
update_data_button = Button(root, text="Update Data", command=update_data)
update_data_button.pack(ipadx=16, pady=10)
updated_status_label = Label(root, textvariable=eft_status, bg="white", fg="#FF0000")
updated_status_label.pack(pady=2)

# Create a button to export to Excel and its status label
export_button = Button(root, text="Export to Excel", command=lambda: export_to_excel(eft_file_df, updated_df), bg="#009688", fg="white", state="normal")
export_button.pack(ipadx=10, pady=5)
export_status_label = Label(root, textvariable=export_status, bg="white", fg="#FF0000")
export_status_label.pack(pady=2)

# Add the 'Create new EFT file' button and its status label
create_eft_button = Button(root, text="Create new EFT file", command=create_new_eft_file, bg="#66FFCC", fg="black")  # Prev color #CC0000
create_eft_button.pack(pady=5)
eft_creation_status_label = Label(root, textvariable=eft_creation_status, bg="white", fg="#FF0000")
eft_creation_status_label.pack(pady=2)

# Run the Tkinter main event loop
root.mainloop()